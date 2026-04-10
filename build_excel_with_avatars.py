"""
Build comprehensive Excel (.xlsx) with avatar images embedded in cells.
Downloads smallHeadURL images and inserts them into the Excel sheet.
"""

import csv
import json
import os
import sys
import io
import urllib.request
import time
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ---- Config ----
INPUT_CSV = 'group_members_comprehensive.csv'
OUTPUT_XLSX = 'group_members_comprehensive.xlsx'
AVATAR_DIR = 'avatars_cache'
AVATAR_SIZE = 40  # px, for Excel cell
ROW_HEIGHT = 35   # Excel row height in points
BASE_API = 'http://127.0.0.1:5200/api/v1'

# Columns to include (reorder: put avatar right after index)
COLUMNS = [
    ('序号', 6),
    ('头像', 7),          # image column
    ('可辨识昵称', 20),
    ('微信昵称', 16),
    ('群昵称', 16),
    ('活跃昵称', 20),
    ('微信ID', 22),
    ('微信号', 16),
    ('备注', 14),
    ('发言总数', 10),
    ('首次发言', 12),
    ('最后发言', 12),
    ('距今天数', 8),
    ('活跃得分', 8),
    ('活跃状态', 10),
    ('清理建议', 12),
    ('是否群主', 8),
]

# ---- 1. Read CSV data ----
print(f'Reading {INPUT_CSV}...')
rows = []
with open(INPUT_CSV, 'r', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f)
    for row in reader:
        rows.append(row)
print(f'  {len(rows)} rows')

# ---- 2. Fill missing avatar URLs via /contacts/:id API ----
missing_avatar_uids = [row['微信ID'] for row in rows if not row.get('头像URL(小)')]
if missing_avatar_uids:
    print(f'\n{len(missing_avatar_uids)} members missing avatar URL, fetching from API...')
    api_fetched = 0
    api_failed = 0
    # Build a lookup for quick row access
    row_by_uid = {row['微信ID']: row for row in rows}
    for i, uid in enumerate(missing_avatar_uids):
        url = f'{BASE_API}/contacts/{uid}'
        try:
            with urllib.request.urlopen(url, timeout=10) as resp:
                cdata = json.loads(resp.read().decode('utf-8'))
            c = cdata.get('data', {}) if isinstance(cdata, dict) else {}
            small_url = c.get('smallHeadImgUrl', '')
            big_url = c.get('bigHeadImgUrl', '')
            if small_url:
                row_by_uid[uid]['头像URL(小)'] = small_url
                row_by_uid[uid]['头像URL(大)'] = big_url
                api_fetched += 1
        except Exception as e:
            api_failed += 1
        if (i + 1) % 50 == 0:
            print(f'  {i+1}/{len(missing_avatar_uids)} queried ({api_fetched} got avatar, {api_failed} failed)')
    print(f'  API done: {api_fetched} new avatars, {api_failed} failed')

# ---- 3. Download avatars ----
os.makedirs(AVATAR_DIR, exist_ok=True)

def download_avatar(url, uid):
    """Download and resize avatar, return path or None."""
    if not url:
        return None
    
    safe_name = uid.replace('@', '_').replace('.', '_')
    path = os.path.join(AVATAR_DIR, f'{safe_name}.png')
    
    # Use cached version if exists
    if os.path.exists(path) and os.path.getsize(path) > 0:
        return path
    
    try:
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = resp.read()
        
        # Save and resize
        tmp_path = path + '.tmp'
        with open(tmp_path, 'wb') as f:
            f.write(data)
        
        img = PILImage.open(tmp_path)
        img = img.resize((AVATAR_SIZE, AVATAR_SIZE), PILImage.LANCZOS)
        img.save(path, 'PNG')
        os.remove(tmp_path)
        return path
    except Exception as e:
        return None

# Collect URLs and download
avatar_urls = {}
for row in rows:
    uid = row['微信ID']
    url = row.get('头像URL(小)', '')
    if url:
        avatar_urls[uid] = url

print(f'Downloading {len(avatar_urls)} avatars (with cache)...')
avatar_paths = {}
failed = 0
for i, (uid, url) in enumerate(avatar_urls.items()):
    path = download_avatar(url, uid)
    if path:
        avatar_paths[uid] = path
    else:
        failed += 1
    
    if (i + 1) % 50 == 0:
        print(f'  {i+1}/{len(avatar_urls)} downloaded ({len(avatar_paths)} ok, {failed} failed)')

print(f'  Done: {len(avatar_paths)} avatars, {failed} failed')

# ---- 3. Build Excel ----
print(f'\nBuilding Excel...')
wb = Workbook()
ws = wb.active
ws.title = 'Group Members'

# Header style
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_white = Font(bold=True, size=11, color='FFFFFF')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

# Write header row
for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = col_width

ws.row_dimensions[1].height = 20

# Color for inactive rows
inactive_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
remove_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')

# Avatar column index (1-based)
avatar_col_idx = 2  # '头像' is 2nd column

# Write data rows
print(f'Writing {len(rows)} rows with images...')
for row_idx, row in enumerate(rows, 2):  # start at row 2 (after header)
    uid = row['微信ID']
    is_inactive = row.get('活跃状态', '') == '从未发言'
    is_remove = row.get('清理建议', '') == '建议移除'
    
    for col_idx, (col_name, _) in enumerate(COLUMNS, 1):
        if col_name == '头像':
            # Will insert image separately
            continue
        
        # Map column name to CSV field
        csv_key = col_name
        value = row.get(csv_key, '')
        
        # Convert numeric fields
        if col_name in ('发言总数', '距今天数') and value:
            try:
                value = int(value)
            except:
                pass
        elif col_name == '活跃得分' and value:
            try:
                value = float(value)
            except:
                pass
        
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = thin_border
        
        if is_remove:
            cell.fill = remove_fill
        elif is_inactive:
            cell.fill = inactive_fill
    
    # Set row height for avatar
    ws.row_dimensions[row_idx].height = ROW_HEIGHT
    
    # Insert avatar image
    avatar_path = avatar_paths.get(uid)
    if avatar_path and os.path.exists(avatar_path):
        try:
            img = XlImage(avatar_path)
            img.width = AVATAR_SIZE
            img.height = AVATAR_SIZE
            # Anchor to the cell
            cell_ref = f'{get_column_letter(avatar_col_idx)}{row_idx}'
            ws.add_image(img, cell_ref)
        except Exception as e:
            pass
    
    if (row_idx - 1) % 100 == 0:
        print(f'  {row_idx - 1}/{len(rows)} rows written')

# Freeze header row
ws.freeze_panes = 'A2'

# Auto-filter
ws.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}'

# ---- 4. Save ----
print(f'Saving {OUTPUT_XLSX}...')
wb.save(OUTPUT_XLSX)
file_size = os.path.getsize(OUTPUT_XLSX) / 1024 / 1024
print(f'\nDone! {OUTPUT_XLSX} ({file_size:.1f} MB)')
print(f'  {len(rows)} members, {len(avatar_paths)} avatars embedded')
print(f'  Row coloring: pink = suggest remove, yellow = never spoke')
